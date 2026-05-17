import win32com.client
import os
import re
import fitz  # PyMuPDF
import sys
import zipfile
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Comprehensive Bijoy ANSI to Unicode Bengali mappings
replacements = {
    # === Vowel signs ===
    'Ï': 'ে',
    'Ì': '',       # hasanta/dead char
    'ï': 'উ',
    'Ĳ': 'উ',
    '◌': '',
    '○': '',

    # === Common conjuncts (যুক্তাক্ষর) ===
    'Ë': '্য',      # -ya
    'Î': 'র্',      # r-hasanta (ref)
    'Ø': 'ক্র',     # kra
    'Ñ': 'ক্ক',     # kka
    'Ù': 'ক্ষ',     # ksha
    'Ô': 'ক্স',     # ksa
    'Ó': 'ক্ত',     # kta
    'Ĕ': 'ত্র',     # tra
    'Ĵ': 'প্র',     # pra
    'Ř': 'শ্র',     # shra
    'Ľ': 'ব্র',     # bra
    'ľ': 'ম্র',     # mra
    'ņ': 'ম্র',     # mra (alt)
    'Ş': 'স্ট',     # stha
    'Š': 'ন্ট',     # nta (retroflex)
    'Ň': 'ম্ম',     # mma
    'Ō': 'ল্ল',     # lla
    'ĺ': 'ব্দ',     # bda
    'õ': 'জ্জ',     # jja
    'Ģ': 'ন্ত',     # nta
    'ē': 'ত্ত',     # tta
    'ł': 'ত্ত',     # tta (alt)
    'Ő': 'ল্প',     # lpa
    'û': 'ঞ্চ',     # ncha
    'ê': 'ঙ্গ',     # ngga
    'ý': 'ঞ্জ',     # nja
    'Œ': 'ঞ্জ',     # nja (alt)
    'š': 'চ্ছ',     # cchha
    'ś': 'শ্ব',     # shba
    'ũ': 'স্ত',     # sta
    'Ǝ': 'স্ত',     # sta (alt)
    'ń': 'ম্ব',     # mba
    'Ĥ': 'ন্দ',     # nda
    'ħ': 'ন্দ্র',   # ndra
    'Ĩ': 'ন্ন',     # nna
    'ĥ': 'ন্ম',     # nma
    'Ŧ': 'স্ক',     # ska
    'î': 'চ্চ',     # ccha
    'ū': 'স্ন',     # sna
    'Ů': 'স্ব',     # sba
    'ă': 'ড্র',     # dra (retroflex)
    'ř': 'শ্য',     # shya
    'ĩ': 'ন্স',     # nsa
    'ű': 'স্ত্র',   # stra
    'Ŗ': 'স্ত্র',   # stra (alt)
    'ů': 'স্ম',     # sma
    'ş': 'ষ্ট',     # shta
    'İ': 'প্ত',     # pta
    'Ţ': 'স্ট্র',   # stra (retroflex)
    'ĳ': 'ন্ধ্য',   # ndhya
    'į': 'ন্য',     # nya
    'ę': 'দ্রি',    # dri
    'ė': 'দ্দী',    # ddi
    'ļ': 'ব্ব',     # bba
    'ƅ': 'হৃ',      # hri
    '\x8c': 'ন্',   # n-hasanta
    '\x95': 'স্',   # s-hasanta
    '\x8e': 'ফ্ফ',  # ffa
    '\x98': 'হ',     # ha
    '×': 'খ',       # kha

    # === Vowel-modified consonants (কারযুক্ত) ===
    'ƃ': 'দু',      # du
    'Ɓ': 'রু',      # ru
    'ƀ': 'সু',      # su
    'ƣ': 'কু',      # ku
    'ſ': 'নু',      # nu
    'Ž': 'গু',      # gu
    'Ƅ': 'শু',      # shu
    'Ɔ': 'হু',      # hu

    # === Standalone / special ===
    'Ð': 'ড',       # da (retroflex)
    'Ə': 'দ',       # da

    # === Additional conjuncts (extended) ===
    'Ã': 'গ্র',     # gra
    'Å': 'ঙ্ক',     # ngka
    'Æ': 'ঙ্খ',     # ngkha
    'Ç': 'জ্ঞ',     # jnya
    'Ú': 'ক্ষ্ম',   # kshma
    'Û': 'গ্ন',     # gna
    'Ü': 'ঘ্ন',     # ghna
    'Þ': 'ণ্ঠ',     # ntha (retroflex)
    'ß': 'ণ্ড',     # nda (retroflex)
    'à': 'ড্ড',     # dda (retroflex)
    'á': 'দ্ধ',     # ddha
    'â': 'দ্ভ',     # dbha
    'å': 'ত্ন',     # tna
    'æ': 'ত্ব',     # tba
    'ç': 'ত্ম',     # tma
    'è': 'দ্গ',     # dga
    'é': 'দ্ঘ',     # dgha
    'ì': 'দ্ব',     # dba
    'í': 'দ্ম',     # dma
    'ð': 'ধ্ব',     # dhba
    'ñ': 'ন্ত্র',   # ntra
    'ò': 'ন্ঠ',     # ntha
    'ó': 'ন্ড',     # nda
    'ö': 'ন্থ',     # ntha
    'ø': 'ন্ধ',     # ndha
    'ù': 'ন্ব',     # nba
    'ú': 'প্ন',     # pna
    'ü': 'ম্ন',     # mna
    'þ': 'প্ল',     # pla
    'ÿ': 'ব্জ',     # bja
    'Ä': 'গ্ল',     # gla
    'È': 'ঞ্ঝ',     # njha
    'É': 'ট্ট',     # tta (retroflex)
    'Ê': 'ট্র',     # tra (retroflex)
    'Ö': 'ক্ল',     # kla
    'Ý': 'চ্ব',     # cba
    'ä': 'ড্ব',     # dba (retroflex)
    'ë': 'ত্থ',     # ttha
    'ì': 'দ্ব',     # dba
    'ô': 'ন্ড্র',   # ndra
}

def clean_text(t):
    if not t: return t
    for k, v in replacements.items():
        t = t.replace(k, v)
        
    cluster = r'([\u0995-\u09B9\u09DC-\u09DF](?:\u09CD[\u0995-\u09B9\u09DC-\u09DF])*)'
    t = re.sub(r'\u09C7' + cluster + r'\u09BE', r'\1ো', t)
    t = re.sub(r'\u09C7' + cluster + r'\u09D7', r'\1ৌ', t)
    t = re.sub(r'\u09C7' + cluster, r'\1ে', t)
    t = re.sub(r'\u09C8' + cluster, r'\1ৈ', t)
    t = re.sub(r'\u09BF' + cluster, r'\1ি', t)
    
    # Apply Bengali spell corrections
    t = bn_spell_fix(t)
    return t

# === Bengali Spell-Check / Correction System ===

# Common misspellings → correct Bengali (voter data specific)
bn_corrections = {
    # --- Occupations (পেশা) ---
    'কষৃি': 'কৃষি',
    'কৃিষ': 'কৃষি',
    'গহৃিণী': 'গৃহিণী',
    'গৃিহণী': 'গৃহিণী',
    'গহৃীণি': 'গৃহিণী',
    'গহৃিনী': 'গৃহিণী',
    'গৃহীণি': 'গৃহিণী',
    'বযব্সা': 'ব্যবসা',
    'ব্যাবসা': 'ব্যবসা',
    'বাযব্সা': 'ব্যবসা',
    'বযবসা': 'ব্যবসা',
    'চাকিরি': 'চাকরি',
    'চাকুরি': 'চাকরি',
    'চাকুরী': 'চাকরি',
    'শমৃিক': 'শ্রমিক',
    'শ্রিমক': 'শ্রমিক',
    'শমৃক': 'শ্রমিক',
    'ছাত্র্': 'ছাত্র',
    'দিনমজরু': 'দিনমজুর',
    'দিনমজরূ': 'দিনমজুর',
    'িদনমজুর': 'দিনমজুর',
    'মৎসযজীবী': 'মৎস্যজীবী',
    'মতস্যজীবী': 'মৎস্যজীবী',
    'শিক্ষাক': 'শিক্ষক',
    'িশক্ষক': 'শিক্ষক',
    'ডাক্তর': 'ডাক্তার',
    'ইিঞ্জিনয়ার': 'ইঞ্জিনিয়ার',
    'ইমাম': 'ইমাম',
    'রকিশাচালক': 'রিকশাচালক',
    'িরকশাচালক': 'রিকশাচালক',
    'অবসরপ্রাপ্ত্': 'অবসরপ্রাপ্ত',
    
    # --- Administrative terms ---
    'ইউিনয়ন': 'ইউনিয়ন',
    'উপজলো': 'উপজেলা',
    'উপেজলা': 'উপজেলা',
    'পৌরসভো': 'পৌরসভা',
    'িসটি': 'সিটি',
    'কর্পোরশেন': 'কর্পোরেশন',
    'কর্পোরশন': 'কর্পোরেশন',
    'ওয়াডর্': 'ওয়ার্ড',
    'ওয়াড': 'ওয়ার্ড',
    'পিরষদ': 'পরিষদ',
    'পিরশদ': 'পরিষদ',
    'ডাকঘড়': 'ডাকঘর',
    
    # --- Common name parts ---
    'মোাঃ': 'মোঃ',
    'মোা:': 'মোঃ',
    'মোসাঃ': 'মোসাঃ',
    'আবুদল': 'আবদুল',
    'আবদলু': 'আবদুল',
    'মোহাম্মাদ': 'মোহাম্মদ',
    'রহামান': 'রহমান',
    'রহামন': 'রহমান',
    'ইসালাম': 'ইসলাম',
    'ইসলমা': 'ইসলাম',
    'হোসনে': 'হোসেন',
    'হোসনে': 'হোসেন',
    'বেগামে': 'বেগম',
    'খাতনু': 'খাতুন',
    
    # --- Field labels (if corrupted) ---
    'ভোটাড়': 'ভোটার',
    'ভোটাার': 'ভোটার',
    'পশো': 'পেশা',
    'পশেো': 'পেশা',
    'ঠিকাানা': 'ঠিকানা',
    'িঠকানা': 'ঠিকানা',
    'জন্ম তািরখ': 'জন্ম তারিখ',
    'তািরখ': 'তারিখ',
    'মাতাা': 'মাতা',
    'িপতা': 'পিতা',
    
    # --- Common place name fixes ---
    'দিনাজপরু': 'দিনাজপুর',
    'িদনাজপুর': 'দিনাজপুর',
    'বাংলাদশে': 'বাংলাদেশ',
    'বালংাদেশ': 'বাংলাদেশ',
    'ঢাকো': 'ঢাকা',
}

def bn_spell_fix(text):
    """Apply Bengali spelling corrections to text."""
    if not text:
        return text
    for wrong, correct in bn_corrections.items():
        text = text.replace(wrong, correct)
    return text

def strip_non_bengali(text):
    """Remove stray non-Bengali characters from a Bengali text field.
    Keeps: Bengali Unicode block, Bengali digits, spaces, punctuation, numbers."""
    if not text:
        return text
    cleaned = []
    for ch in text:
        cp = ord(ch)
        if (0x0980 <= cp <= 0x09FF or       # Bengali Unicode block
            ch in ' \t।,:;.-/()' or          # Allowed punctuation
            '0' <= ch <= '9' or              # ASCII digits
            '০' <= ch <= '৯'):               # Bengali digits
            cleaned.append(ch)
    return ''.join(cleaned).strip()


def parse_date(date_str):
    """Convert Bengali format like ০১/০১/১৯৯০ or 01/01/1990 to YYYY-MM-DD"""
    if not date_str:
        return ""
    # Convert Bengali numerals to English numerals
    b_to_e = {'০':'0', '১':'1', '২':'2', '৩':'3', '৪':'4', '৫':'5', '৬':'6', '৭':'7', '৮':'8', '৯':'9'}
    for b, e in b_to_e.items():
        date_str = date_str.replace(b, e)
    
    # Try to extract DD/MM/YYYY
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', date_str)
    if m:
        dd = m.group(1).zfill(2)
        mm = m.group(2).zfill(2)
        yyyy = m.group(3)
        return f"{yyyy}-{mm}-{dd}"
    return date_str

def parse_text(text):
    lines = text.splitlines()
    
    meta = {
        'division_name': '',
        'district_name': '',
        'upazila_name': '',
        'pourosova_name': '',
        'union_name': '',
        'word_number': '',
        'alaka_name': '',
        'alaka_number': '',
        'post_office': '',
        'post_code': '',
        'city_corporation': '',
        'ason_number': '',
        'global_gender': ''
    }
    
    records = []
    current_record = {}
    prev_line = ""
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        c_line = clean_text(line).strip()
        
        # State machine for next-line headers
        if 'উপজেলা/থানা' in prev_line and not meta['upazila_name']:
            meta['upazila_name'] = c_line
        elif ('ইউনিয়ন/ পৌর ওয়ার্ড' in prev_line or 'ক্যান্টনমেন্ট বো' in prev_line) and not meta['union_name']:
            meta['union_name'] = c_line
        elif 'ভোটার এলাকা' == prev_line and not meta['alaka_name']:
            meta['alaka_name'] = c_line
        elif 'ভোটার এলাকার নম্বর' == prev_line and not meta['alaka_number']:
            meta['alaka_number'] = c_line
        elif ('সিটি কর্পোরেশন' in prev_line or 'পৌরসভা' in prev_line) and not meta['pourosova_name']:
            if 'নম্বর (' not in c_line and 'ওয়াড' not in c_line:
                meta['pourosova_name'] = c_line
                meta['city_corporation'] = c_line
        elif ('ওয়ার্ড নম্বর' in prev_line or 'ওয়াডর্ নম্বর' in prev_line) and not meta['word_number']:
            meta['word_number'] = c_line
            
        # Global Gender
        if '(পুরুষ)' in c_line:
            meta['global_gender'] = 'পুরুষ'
        elif '(মহিলা)' in c_line:
            meta['global_gender'] = 'মহিলা'
            
        # Colon separated headers
        if c_line.startswith('অঞ্চল:') and not meta['division_name']:
            meta['division_name'] = c_line.split(':', 1)[1].strip()
        elif c_line.startswith('জেলা:') and not meta['district_name']:
            meta['district_name'] = c_line.split(':', 1)[1].strip()
        elif (c_line.startswith('পোস্টকোড') or c_line.startswith('পোস্ট কোড')) and not meta['post_code']:
            parts = c_line.split(':')
            if len(parts) > 1:
                meta['post_code'] = parts[1].strip()
        elif c_line.startswith('ডাকঘর:') and not meta['post_office']:
            meta['post_office'] = c_line.split(':', 1)[1].strip()
        elif c_line.startswith('আসন:') and not meta['ason_number']:
            meta['ason_number'] = c_line.split(':', 1)[1].strip()

        # Match lines like "০০১. নাম: মোঃ ওবাইদুর রহমান"
        m_name = re.match(r'^([০-৯0-9]+)\.\s*নাম:\s*(.*)', c_line)
        if m_name:
            if current_record:
                records.append(current_record)
            current_record = {
                'name': m_name.group(2),
                'voter_number': '',
                'father_name': '',
                'mother_name': '',
                'ocapation': '',
                'dob': '',
                'address': '',
                'gender': meta.get('global_gender', '')
            }
            prev_line = c_line
            continue
            
        if current_record:
            if c_line.startswith('ভোটার নং:'):
                current_record['voter_number'] = c_line.replace('ভোটার নং:', '').strip()
            elif c_line.startswith('পিতা:'):
                current_record['father_name'] = c_line.replace('পিতা:', '').strip()
            elif c_line.startswith('স্বামী:'):
                current_record['father_name'] = c_line.replace('স্বামী:', '').strip()
                if not current_record['gender']: current_record['gender'] = 'মহিলা'
            elif c_line.startswith('মাতা:'):
                current_record['mother_name'] = c_line.replace('মাতা:', '').strip()
            elif c_line.startswith('পেশা:'):
                parts = c_line.split(',জন্ম তারিখ:')
                current_record['ocapation'] = parts[0].replace('পেশা:', '').strip()
                if len(parts) > 1:
                    current_record['dob'] = parts[1].strip()
            elif c_line.startswith('ঠিকানা:'):
                current_record['address'] = c_line.replace('ঠিকানা:', '').strip()
                
        prev_line = c_line
        
    if current_record:
        records.append(current_record)
        
    return meta, records

def get_output_filename(input_dir):
    try:
        for item in os.listdir(input_dir):
            item_path = os.path.join(input_dir, item)
            if os.path.isdir(item_path):
                return item
    except Exception:
        pass
    return "ExportedDatabase"

def escape_sql(val):
    if not val:
        return "NULL"
    val_str = str(val).replace("'", "''")
    return f"'{val_str}'"

def extract_numbers_only(val):
    if not val:
        return ""
    b_to_e = {'০':'0', '১':'1', '২':'2', '৩':'3', '৪':'4', '৫':'5', '৬':'6', '৭':'7', '৮':'8', '৯':'9'}
    for b, e in b_to_e.items(): val = val.replace(b, e)
    return re.sub(r'[^0-9]', '', val)

def main():
    print("=" * 50)
    print("    OCR Data Extraction Automation (ZIP Mode)")
    print("=" * 50)
    print("Select Export Format:")
    print("1. .sql  (Fast, for web/MySQL servers)")
    print("2. .accdb (Microsoft Access)")
    print("3. .xlsx (Microsoft Excel)")
    choice = input("\nEnter 1, 2 or 3: ").strip()
    
    if choice not in ('1', '2', '3'):
        print("Invalid choice. Exiting.")
        return

    root_dir = os.getcwd()
    input_dir = os.path.join(root_dir, 'Input')
    output_dir = os.path.join(root_dir, 'Output')
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    if not os.path.exists(input_dir):
        print("\nError: 'Input' directory not found.")
        return
    
    # Find all ZIP files in Input/
    zip_files = sorted([f for f in os.listdir(input_dir) if f.lower().endswith('.zip')])
    
    if not zip_files:
        print("No ZIP files found in Input directory.")
        return
    
    print(f"\nFound {len(zip_files)} ZIP file(s):")
    for zf in zip_files:
        print(f"  • {zf}")
    print()
    
    grand_total_records = 0
    grand_total_pdfs = 0
    grand_success_zips = 0
    grand_fail_zips = 0
    
    for zip_idx, zip_filename in enumerate(zip_files, 1):
        zip_path = os.path.join(input_dir, zip_filename)
        
        # Extract ason_number from filename: "example-১.zip" -> "১"
        base_name = os.path.splitext(zip_filename)[0]
        ason_number = base_name.rsplit('-', 1)[1].strip() if '-' in base_name else ''
        
        ext = {'1': '.sql', '2': '.accdb', '3': '.xlsx'}[choice]
        output_file = os.path.join(output_dir, f"{base_name}{ext}")
        
        print(f"{'=' * 50}")
        print(f"[ZIP {zip_idx}/{len(zip_files)}] {zip_filename}")
        print(f"  ason_number : {ason_number}")
        print(f"  Output      : {base_name}{ext}")
        print(f"{'=' * 50}")
        
        success_count = 0
        fail_count = 0
        total_records_inserted = 0
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                pdf_names = sorted([n for n in z.namelist() if n.lower().endswith('.pdf')])
                total_pdfs = len(pdf_names)
                
                if total_pdfs == 0:
                    print(f"  No PDFs found inside. Skipping.\n")
                    grand_fail_zips += 1
                    continue
                
                print(f"  PDFs inside : {total_pdfs}\n")
                
                if choice == '1':
                    # ========== SQL EXPORT ==========
                    try:
                        with open(output_file, 'w', encoding='utf-8-sig') as f:
                            f.write("-- Database Export Generated by OCR Automation\n")
                            f.write(f"-- Source: {zip_filename}\n")
                            f.write(f"-- Ason Number: {ason_number}\n\n")
                            f.write("CREATE TABLE IF NOT EXISTS `Voters` (\n")
                            f.write("  `division_name` VARCHAR(200),\n")
                            f.write("  `district_name` VARCHAR(200),\n")
                            f.write("  `upazila_name` VARCHAR(200),\n")
                            f.write("  `pourosova_name` VARCHAR(200),\n")
                            f.write("  `union_name` VARCHAR(200),\n")
                            f.write("  `name` VARCHAR(150),\n")
                            f.write("  `father_name` VARCHAR(150),\n")
                            f.write("  `mother_name` VARCHAR(150),\n")
                            f.write("  `ocapation` VARCHAR(200),\n")
                            f.write("  `dob` VARCHAR(50),\n")
                            f.write("  `address` VARCHAR(300),\n")
                            f.write("  `gender` VARCHAR(200),\n")
                            f.write("  `word_number` VARCHAR(50),\n")
                            f.write("  `alaka_name` VARCHAR(200),\n")
                            f.write("  `alaka_number` VARCHAR(50),\n")
                            f.write("  `post_office` VARCHAR(200),\n")
                            f.write("  `post_code` VARCHAR(50),\n")
                            f.write("  `voter_number` VARCHAR(200),\n")
                            f.write("  `city_corporation` VARCHAR(200),\n")
                            f.write("  `ason_number` VARCHAR(200)\n")
                            f.write(");\n\n")

                            for pdf_name in pdf_names:
                                try:
                                    pdf_bytes = z.read(pdf_name)
                                    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                                    text = ""
                                    for page in doc:
                                        text += page.get_text()
                                    doc.close()
                                    
                                    meta_clean, records = parse_text(text)
                                    meta_clean['ason_number'] = ason_number
                                    
                                    if records:
                                        short_name = os.path.basename(pdf_name)
                                        f.write(f"\n-- Data for {short_name}\n")
                                        f.write("INSERT INTO `Voters` (`division_name`, `district_name`, `upazila_name`, `pourosova_name`, `union_name`, `name`, `father_name`, `mother_name`, `ocapation`, `dob`, `address`, `gender`, `word_number`, `alaka_name`, `alaka_number`, `post_office`, `post_code`, `voter_number`, `city_corporation`, `ason_number`) VALUES \n")
                                        
                                        val_strings = []
                                        for r in records:
                                            try:
                                                d_name = escape_sql(meta_clean.get('division_name', ''))
                                                dist_name = escape_sql(meta_clean.get('district_name', ''))
                                                upa_name = escape_sql(meta_clean.get('upazila_name', ''))
                                                poro_name = escape_sql(meta_clean.get('pourosova_name', ''))
                                                uni_name = escape_sql(meta_clean.get('union_name', ''))
                                                name = escape_sql(r.get('name', ''))
                                                fname = escape_sql(r.get('father_name', ''))
                                                mname = escape_sql(r.get('mother_name', ''))
                                                prof = escape_sql(r.get('ocapation', ''))
                                                
                                                dob_bn = r.get('dob', '')
                                                dob = escape_sql(dob_bn) if dob_bn else "NULL"
                                                
                                                addr = escape_sql(r.get('address', ''))
                                                gender = escape_sql(r.get('gender', ''))
                                                
                                                ward_raw = meta_clean.get('word_number', '')
                                                ward_val = escape_sql(ward_raw) if ward_raw else "NULL"
                                                
                                                alaka_name_val = escape_sql(meta_clean.get('alaka_name', ''))
                                                
                                                alaka_raw = meta_clean.get('alaka_number', '')
                                                alaka_num_val = escape_sql(alaka_raw) if alaka_raw else "NULL"
                                                
                                                post_off = escape_sql(meta_clean.get('post_office', ''))
                                                
                                                pc_raw = meta_clean.get('post_code', '')
                                                pc_val = escape_sql(pc_raw) if pc_raw else "NULL"
                                                
                                                vid_raw = r.get('voter_number', '')
                                                vid_val = escape_sql(vid_raw) if vid_raw else "NULL"
                                                
                                                city_corp = escape_sql(meta_clean.get('city_corporation', ''))
                                                ason = escape_sql(meta_clean.get('ason_number', ''))
                                                
                                                val_strings.append(f"({d_name}, {dist_name}, {upa_name}, {poro_name}, {uni_name}, {name}, {fname}, {mname}, {prof}, {dob}, {addr}, {gender}, {ward_val}, {alaka_name_val}, {alaka_num_val}, {post_off}, {pc_val}, {vid_val}, {city_corp}, {ason})")
                                                total_records_inserted += 1
                                            except Exception as row_err:
                                                print(f"  [WARNING] Row error in {os.path.basename(pdf_name)}: {row_err}")
                                            
                                        if val_strings:
                                            f.write(",\n".join(val_strings) + ";\n")
                                        
                                    success_count += 1
                                    print(f"  [{success_count+fail_count}/{total_pdfs}] {os.path.basename(pdf_name)}")
                                except Exception as e:
                                    print(f"  [ERROR] {os.path.basename(pdf_name)}: {e}")
                                    fail_count += 1
                                    
                    except Exception as e:
                        print(f"  [ERROR] Failed to write SQL: {e}")
                        grand_fail_zips += 1
                        continue

                elif choice == '2':
                    # ========== ACCDB EXPORT ==========
                    if os.path.exists(output_file):
                        try:
                            os.remove(output_file)
                        except PermissionError:
                            print(f"  [ERROR] Cannot overwrite '{os.path.basename(output_file)}'. Close it in Access first.")
                            grand_fail_zips += 1
                            continue

                    conn_str = f"Provider=Microsoft.ACE.OLEDB.12.0;Data Source={output_file};"
                    try:
                        catalog = win32com.client.Dispatch("ADOX.Catalog")
                        catalog.Create(conn_str)
                    except Exception as e:
                        print(f"  [ERROR] Failed to create Access DB: {e}")
                        grand_fail_zips += 1
                        continue

                    conn = None
                    try:
                        conn = win32com.client.Dispatch("ADODB.Connection")
                        conn.Open(conn_str)

                        create_table_sql = """
                        CREATE TABLE Voters (
                            division_name VARCHAR(200),
                            district_name VARCHAR(200),
                            upazila_name VARCHAR(200),
                            pourosova_name VARCHAR(200),
                            union_name VARCHAR(200),
                            name VARCHAR(150),
                            father_name VARCHAR(150),
                            mother_name VARCHAR(150),
                            ocapation VARCHAR(200),
                            dob VARCHAR(50),
                            address VARCHAR(255),
                            gender VARCHAR(200),
                            word_number VARCHAR(50),
                            alaka_name VARCHAR(200),
                            alaka_number VARCHAR(50),
                            post_office VARCHAR(200),
                            post_code VARCHAR(50),
                            voter_number VARCHAR(200),
                            city_corporation VARCHAR(200),
                            ason_number VARCHAR(200)
                        )
                        """
                        conn.Execute(create_table_sql)

                        rs = win32com.client.Dispatch("ADODB.Recordset")
                        rs.Open("Voters", conn, 1, 3)

                        for pdf_name in pdf_names:
                            try:
                                pdf_bytes = z.read(pdf_name)
                                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                                text = ""
                                for page in doc:
                                    text += page.get_text()
                                doc.close()
                                
                                meta_clean, records = parse_text(text)
                                meta_clean['ason_number'] = ason_number
                                
                                for r in records:
                                    try:
                                        rs.AddNew()
                                        rs.Fields.Item('division_name').Value = meta_clean.get('division_name', '')
                                        rs.Fields.Item('district_name').Value = meta_clean.get('district_name', '')
                                        rs.Fields.Item('upazila_name').Value = meta_clean.get('upazila_name', '')
                                        rs.Fields.Item('pourosova_name').Value = meta_clean.get('pourosova_name', '')
                                        rs.Fields.Item('union_name').Value = meta_clean.get('union_name', '')
                                        rs.Fields.Item('name').Value = r.get('name', '')
                                        rs.Fields.Item('father_name').Value = r.get('father_name', '')
                                        rs.Fields.Item('mother_name').Value = r.get('mother_name', '')
                                        rs.Fields.Item('ocapation').Value = r.get('ocapation', '')
                                        rs.Fields.Item('dob').Value = r.get('dob', '')
                                        rs.Fields.Item('address').Value = r.get('address', '')
                                        rs.Fields.Item('gender').Value = r.get('gender', '')
                                        rs.Fields.Item('word_number').Value = meta_clean.get('word_number', '')
                                        rs.Fields.Item('alaka_name').Value = meta_clean.get('alaka_name', '')
                                        rs.Fields.Item('alaka_number').Value = meta_clean.get('alaka_number', '')
                                        rs.Fields.Item('post_office').Value = meta_clean.get('post_office', '')
                                        rs.Fields.Item('post_code').Value = meta_clean.get('post_code', '')
                                        rs.Fields.Item('voter_number').Value = r.get('voter_number', '')
                                        rs.Fields.Item('city_corporation').Value = meta_clean.get('city_corporation', '')
                                        rs.Fields.Item('ason_number').Value = meta_clean.get('ason_number', '')
                                        
                                        rs.Update()
                                        total_records_inserted += 1
                                    except Exception as row_err:
                                        print(f"  [WARNING] Row error in {os.path.basename(pdf_name)}: {row_err}")
                                        if rs.EditMode != 0:
                                            rs.CancelUpdate()

                                success_count += 1
                                print(f"  [{success_count+fail_count}/{total_pdfs}] {os.path.basename(pdf_name)}")
                            except Exception as e:
                                print(f"  [ERROR] {os.path.basename(pdf_name)}: {e}")
                                fail_count += 1

                        rs.Close()
                    except Exception as e:
                        print(f"  [ERROR] Database operation failed: {e}")
                        grand_fail_zips += 1
                        continue
                    finally:
                        if conn and conn.State == 1:
                            conn.Close()

                elif choice == '3':
                    # ========== XLSX EXPORT ==========
                    try:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Voters"
                        
                        # Define headers
                        headers = [
                            'division_name', 'district_name', 'upazila_name',
                            'pourosova_name', 'union_name', 'name',
                            'father_name', 'mother_name', 'ocapation',
                            'dob', 'address', 'gender', 'word_number',
                            'alaka_name', 'alaka_number', 'post_office',
                            'post_code', 'voter_number', 'city_corporation',
                            'ason_number'
                        ]
                        
                        # Style header row
                        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
                        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col_idx, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        
                        row_num = 2  # Start data from row 2
                        
                        for pdf_name in pdf_names:
                            try:
                                pdf_bytes = z.read(pdf_name)
                                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                                text = ""
                                for page in doc:
                                    text += page.get_text()
                                doc.close()
                                
                                meta_clean, records = parse_text(text)
                                meta_clean['ason_number'] = ason_number
                                
                                for r in records:
                                    try:
                                        row_data = [
                                            meta_clean.get('division_name', ''),
                                            meta_clean.get('district_name', ''),
                                            meta_clean.get('upazila_name', ''),
                                            meta_clean.get('pourosova_name', ''),
                                            meta_clean.get('union_name', ''),
                                            r.get('name', ''),
                                            r.get('father_name', ''),
                                            r.get('mother_name', ''),
                                            r.get('ocapation', ''),
                                            r.get('dob', ''),
                                            r.get('address', ''),
                                            r.get('gender', ''),
                                            meta_clean.get('word_number', ''),
                                            meta_clean.get('alaka_name', ''),
                                            meta_clean.get('alaka_number', ''),
                                            meta_clean.get('post_office', ''),
                                            meta_clean.get('post_code', ''),
                                            r.get('voter_number', ''),
                                            meta_clean.get('city_corporation', ''),
                                            meta_clean.get('ason_number', '')
                                        ]
                                        for col_idx, val in enumerate(row_data, 1):
                                            cell = ws.cell(row=row_num, column=col_idx, value=val)
                                            cell.border = thin_border
                                        row_num += 1
                                        total_records_inserted += 1
                                    except Exception as row_err:
                                        print(f"  [WARNING] Row error in {os.path.basename(pdf_name)}: {row_err}")
                                
                                success_count += 1
                                print(f"  [{success_count+fail_count}/{total_pdfs}] {os.path.basename(pdf_name)}")
                            except Exception as e:
                                print(f"  [ERROR] {os.path.basename(pdf_name)}: {e}")
                                fail_count += 1
                        
                        # Auto-fit column widths
                        for col_idx, header in enumerate(headers, 1):
                            max_len = len(header)
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    if cell.value:
                                        max_len = max(max_len, len(str(cell.value)))
                            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)
                        
                        # Freeze header row
                        ws.freeze_panes = 'A2'
                        
                        wb.save(output_file)
                        
                    except Exception as e:
                        print(f"  [ERROR] Failed to write XLSX: {e}")
                        grand_fail_zips += 1
                        continue

        except zipfile.BadZipFile:
            print(f"  [ERROR] '{zip_filename}' is not a valid ZIP file. Skipping.")
            grand_fail_zips += 1
            continue
        except Exception as e:
            print(f"  [ERROR] Failed to open {zip_filename}: {e}")
            grand_fail_zips += 1
            continue
        
        grand_total_records += total_records_inserted
        grand_total_pdfs += total_pdfs
        grand_success_zips += 1
        
        status = "✓" if fail_count == 0 else "⚠"
        print(f"\n  {status} Done: {success_count} PDFs, {total_records_inserted} records → {base_name}{ext}\n")

    # Grand Summary
    print("=" * 50)
    print("              GRAND SUMMARY")
    print("=" * 50)
    print(f"  ZIP files processed  : {grand_success_zips}/{len(zip_files)}")
    print(f"  ZIP files failed     : {grand_fail_zips}")
    print(f"  Total PDFs processed : {grand_total_pdfs}")
    print(f"  Total Records        : {grand_total_records}")
    print("=" * 50)
    if grand_fail_zips == 0:
        print("  ✓ All exports completed successfully!")
    else:
        print(f"  ⚠ {grand_fail_zips} ZIP(s) had errors.")
    print("=" * 50)

if __name__ == "__main__":
    main()

